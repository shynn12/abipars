from selenium import webdriver
from openpyxl.styles import PatternFill
from selenium.webdriver.chrome.service import Service
import time
from openpyxl import load_workbook

alf = '0123456789'
fn = 'itmo.xlsx'
base = load_workbook(fn)
ws = base['data']

service = Service(executable_path='C:/chromedriver/chromedriver')
browser = webdriver.Chrome(service=service)


def tabl(url, col, limit):
    browser.get(url)
    time.sleep(1)
    pars = browser.page_source
    row = 3
    tek = 0
    for i in range(len(pars)):
        k = ''
        if pars[i] == '№':
            t = i + 7
            while pars[t] != '<':
                k += pars[t]
                t += 1
            ws.cell(row=row, column=col, value=k)
            tek += 1
            if tek == limit:
                break
            row += 1
    col += 1
    row = 3
    tek = 0
    for i in range(len(pars) - 3):
        k = ''
        if pars[i: i + 3] == 'ада':
            if pars[i - 8] == 'p':
                ws.cell(row=row, column=col, value='ОЛИМПИАДА')
                row += 1
                tek += 1
                if tek == limit:
                    break
                continue
        if pars[i: i + 3] == '+ИД':
            t = i + 11
            while pars[t] != '<':
                k += pars[t]
                t += 1
            if k[0] not in alf:
                break
            ws.cell(row=row, column=col, value=k)
            tek += 1
            if tek == limit:
                break
            row += 1
    col += 1
    row = 3
    tek = 0
    for i in range(len(pars) - 4):
        k = ''
        if pars[i: i + 4] == 'тет:':
            t = i + 11
            while pars[t] != '<':
                k += pars[t]
                t += 1
            if k[0] not in alf:
                break
            ws.cell(row=row, column=col, value=k)
            tek += 1
            if tek == limit:
                break
            row += 1
    base.save(fn)

fillgood = PatternFill('solid',fgColor='008000')
fillbad = PatternFill('solid',fgColor='ff033e')

def anlz(snls, row, col):

    if ws.cell(row=row, column=col + 2).value == '1':
        ws.cell(row=row,column=col).fill = fillgood
        ws.cell(row=row, column=col + 1).fill = fillgood
        ws.cell(row=row, column=col + 2).fill = fillgood
    #else:
    #    for b in range(2, 122 + 1, 5):
    #        a = 3
    #        flag = False
    #        while ws.cell(row=a, column=b).value != None:
    #            if ws.cell(row=a, column=b) == snls:
    #                print(row, col, a, b )
    #                if int(ws.cell(row=a, column=b + 2).value) > int(ws.cell(row=row, column=col + 2).value):
    #                    print('!!!!!!!!!')
    #                    ws.cell(row=a, column=b).fill = fillgood
    #                    ws.cell(row=a, column=b + 1).fill = fillgood
    #                    ws.cell(row=a, column=b + 2).fill = fillgood
    #                    ws.cell(row=row, column=col).fill = fillbad
    #                    ws.cell(row=row, column=col + 1).fill = fillbad
    #                    ws.cell(row=row, column=col + 2).fill = fillbad
    #                    flag = True
    #                    break
    #            a += 1
    #        if flag is True:
    #            break
    #    else:
    #        ws.cell(row=row, column=col).fill = fillgood
    #        ws.cell(row=row, column=col + 1).fill = fillgood
    #        ws.cell(row=row, column=col + 2).fill = fillgood





f = open('sources')
n = int(f.readline())
for i in range(n):
    src, limit = f.readline().split()
    tabl(src, 2 + 5 * i, int(limit))
f.close()
browser.quit()

for b in range(2, 122 + 1, 5):
    a = 3
    while ws.cell(row = a, column = b).value != None:
        anlz(ws.cell(row = a, column= b), a, b)
        a += 1
base.save(fn)
